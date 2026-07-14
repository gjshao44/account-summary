import streamlit as st
import pandas as pd
import glob
import os
import re
import altair as alt
from datetime import datetime
import openpyxl
import subprocess
import sys
import faulthandler

# Enable C-level crash tracing for the main Streamlit process
faulthandler.enable()

# Configure the page first before rendering any elements
st.set_page_config(
    page_title="Portfolio & Budget Analytics Dashboard",
    page_icon="💼",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for strict left-alignment of elements and styling of metric cards
st.markdown("""
    <style>
    /* Pin all Altair charts to the left edge of their containers */
    div[data-testid="stAltairChart"] {
        text-align: left !important;
        margin-left: 0 !important;
        display: block !important;
    }
    
    /* Global metric design refinements */
    div[data-testid="stMetricValue"] {
        font-size: 2rem !important;
        font-weight: 700 !important;
    }
    
    /* Make custom styling containers for alerts */
    .alert-card {
        padding: 1rem;
        border-radius: 0.5rem;
        margin-bottom: 1rem;
        border-left: 5px solid;
    }
    .alert-warning {
        background-color: rgba(255, 193, 7, 0.1);
        border-left-color: #ffc107;
    }
    .alert-danger {
        background-color: rgba(220, 53, 69, 0.1);
        border-left-color: #dc3545;
    }
    .alert-info {
        background-color: rgba(13, 110, 253, 0.1);
        border-left-color: #0d6efd;
    }
    </style>
""", unsafe_allow_html=True)

# Define project directories matching the backends
FINANCIAL_DATA_INPUT = 'financial_data/input'
FINANCIAL_DATA_OUTPUT = 'financial_data/output'
EXPENSE_DATA_INPUT = 'expense_data/input'
EXPENSE_DATA_OUTPUT = 'expense_data/output'



def trigger_portfolio_pipeline():
    """Runs the investment pipeline backend in an isolated subprocess, capturing segmentation faults cleanly."""
    try:
        result = subprocess.run(
            [sys.executable, "-X", "faulthandler", "portfolio_backend.py"],
            capture_output=True,
            text=True,
            check=True
        )
        return True, "Portfolio analytics pipeline finished successfully!"
    except subprocess.CalledProcessError as e:
        # Check specifically for Segmentation Fault exit codes (Unix/Linux/Mac)
        if e.returncode in [-11, 139]:
            return False, f"🚨 **SEGMENTATION FAULT DETECTED IN BACKEND** 🚨\n\nStack Trace:\n{e.stderr}"
        
        error_msg = e.stderr if e.stderr else e.stdout
        return False, f"Portfolio Pipeline process disrupted:\n{error_msg}"
    except Exception as e:
        return False, f"Portfolio Pipeline process disrupted: {e}"

def trigger_expense_pipeline():
    """Runs the expense/budget spreadsheet aggregation backend."""
    try:
        result = subprocess.run(
            [sys.executable, "-X", "faulthandler", "expense_budget_backend.py"],
            capture_output=True,
            text=True,
            check=True
        )
        return True, "Expense aggregation pipeline finished successfully!"
    except subprocess.CalledProcessError as e:
        if e.returncode in [-11, 139]:
            return False, f"🚨 **SEGMENTATION FAULT DETECTED IN BACKEND** 🚨\n\nStack Trace:\n{e.stderr}"
            
        error_msg = e.stderr if e.stderr else e.stdout
        return False, f"Expense Pipeline process disrupted (Code {e.returncode}):\n{error_msg}"
    except Exception as e:
        return False, f"Expense Pipeline process disrupted: {e}"


# STREAMING_CHUNK: Managing automatic execution schedules...

def auto_run_pipelines_if_needed():
    """
    Checks if input files exist but output summaries do not exist.
    If so, runs the workflows automatically so the user doesn't have to press a button.
    """
    # Collect all possible portfolio input files
    portfolio_inputs = (
        glob.glob(os.path.join(FINANCIAL_DATA_INPUT, 'All-Accounts-Positions-*')) +
        glob.glob(os.path.join(FINANCIAL_DATA_INPUT, 'XXXXInvestmentIncome')) +
        glob.glob(os.path.join(FINANCIAL_DATA_INPUT, 'Portfolio_Positions*')) 
    )
    portfolio_outputs = glob.glob(os.path.join(FINANCIAL_DATA_OUTPUT, 'Master_Account_Summary_*.csv'))
    
    # Check Expense
    expense_inputs = glob.glob(os.path.join(EXPENSE_DATA_INPUT, '*transactions*.csv')) + glob.glob(os.path.join(EXPENSE_DATA_INPUT, 'Yearly_budget.xlsx'))
    expense_outputs = glob.glob(os.path.join(EXPENSE_DATA_OUTPUT, 'Yearly_budget.xlsx'))
    
    status_msg = []
    
    if portfolio_inputs and not portfolio_outputs:
        with st.spinner("Processing investment files automatically..."):
            success, msg = trigger_portfolio_pipeline()
            if success:
                status_msg.append("🔄 Auto-generated fresh Portfolio summary.")
                
    if expense_inputs and not expense_outputs:
        with st.spinner("Processing expense sheets automatically..."):
            success, msg = trigger_expense_pipeline()
            if success:
                status_msg.append("📊 Auto-generated fresh Budget and Expense tracking workbook.")
                
    if status_msg:
        st.toast(" | ".join(status_msg), icon="⚡")

# Auto-execute pipelines behind the scenes if raw data is ready
# auto_run_pipelines_if_needed()


# STREAMING_CHUNK: Reading and parsing backend generated alert sheets...

def parse_alerts_log():
    """
    Parses the generated alerts.txt file from expense output folder
    to display category warnings and big transactions on the frontend.
    """
    alert_files = glob.glob(os.path.join(EXPENSE_DATA_OUTPUT, '*_alerts.txt'))
    if not alert_files:
        return None
    
    latest_alert_file = max(alert_files, key=os.path.getmtime)
    
    unmapped_categories = []
    others_rolls = []
    abnormal_txns = []
    
    current_section = None
    
    try:
        with open(latest_alert_file, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                
                # Identify section headings
                if "UNMAPPED CATEGORIES" in line:
                    current_section = "unmapped"
                    continue
                elif "CATEGORIES ROLLED INTO 'OTHERS'" in line:
                    current_section = "others"
                    continue
                elif "ABNORMAL TRANSACTIONS" in line:
                    current_section = "abnormal"
                    continue
                elif line.startswith("===") or line.startswith("Alerts -"):
                    continue
                
                # Parse section items
                if current_section == "unmapped" and line.startswith("-"):
                    unmapped_categories.append(line[2:])
                elif current_section == "others" and line.startswith("-"):
                    others_rolls.append(line[2:])
                elif current_section == "abnormal" and line.startswith("-"):
                    abnormal_txns.append(line[2:])
    except Exception as e:
        return {"error": f"Failed to parse alerts file: {e}"}
        
    return {
        "unmapped": unmapped_categories,
        "others": others_rolls,
        "abnormal": abnormal_txns,
        "file_date": os.path.basename(latest_alert_file).replace('_alerts.txt', '')
    }


# STREAMING_CHUNK: Loading comparative multi-year budget spreadsheets...

def load_historical_budget_data():
    """
    Dynamically loads and parses Yearly_budget.xlsx from the output folder.
    Returns: DataFrames for Expenses and Income, a list of years tracked, and total metrics.
    """
    output_path = os.path.join(EXPENSE_DATA_OUTPUT, 'Yearly_budget.xlsx')
    # Fallback to input if output hasn't been generated yet
    if not os.path.exists(output_path):
        output_path = os.path.join(EXPENSE_DATA_INPUT, 'Yearly_budget.xlsx')
        if not os.path.exists(output_path):
            return None, None, [], {}, {}
            
    try:
        wb = openpyxl.load_workbook(output_path, data_only=True)
    except Exception:
        return None, None, [], {}, {}
    
    def parse_sheet_data(sheet_name):
        if sheet_name not in wb.sheetnames:
            return None, {}, []
            
        ws = wb[sheet_name]
        
        # Determine columns containing years
        col_to_year = {}
        for col in range(2, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val is not None and str(val).strip().isdigit():
                col_to_year[col] = int(str(val).strip())
                
        years = sorted(list(set(col_to_year.values())))
        # Grab up to the last 5 years of history
        historical_years = years[-5:] if len(years) >= 5 else years
        
        # Locate the total row
        total_row_idx = None
        for r in range(1, ws.max_row + 1):
            val = ws.cell(row=r, column=1).value
            if val and str(val).strip().lower() == 'total':
                total_row_idx = r
                break
                
        categories_data = []
        totals_map = {yr: 0.0 for yr in historical_years}
        
        limit_row = total_row_idx if total_row_idx else ws.max_row + 1
        for r in range(2, limit_row):
            cat = ws.cell(row=r, column=1).value
            if cat is None or not str(cat).strip():
                continue
            cat = str(cat).strip()
            
            row_dict = {'Category': cat}
            for col, yr in col_to_year.items():
                if yr in historical_years:
                    val = ws.cell(row=r, column=col).value
                    row_dict[yr] = float(val) if val is not None else 0.0
            categories_data.append(row_dict)
            
        # Dynamically calculate sheet totals in Python memory to prevent None errors before Excel executes
        for yr in historical_years:
            totals_map[yr] = sum(row[yr] for row in categories_data)
        
        # Convert all categories list to a DataFrame
        df = pd.DataFrame(categories_data)
        
        # Cast headers to strings to resolve PyArrow warnings
        df.columns = df.columns.astype(str)
                
        return df, totals_map, historical_years

    df_exp, tot_exp, years_exp = parse_sheet_data('Expense')
    df_inc, tot_inc, years_inc = parse_sheet_data('Income')
    
    # Reconcile years across sheets
    all_years = sorted(list(set(years_exp + years_inc)))
    
    return df_exp, df_inc, all_years, tot_exp, tot_inc


# STREAMING_CHUNK: Styling the sidebar controls...

# Sidebar branding and triggers
with st.sidebar:
    st.image("https://img.icons8.com/fluent/96/000000/money-bag.png", width=64)
    st.title("Sync & Refresh")
    st.write("Update your account and budget data")
    
    st.markdown("---")
    
    # Portfolio Control
    st.subheader("💼 Investment Statements")
    if st.button("🔄 Sync Brokerage Data"):
        with st.spinner("Processing investment files..."):
            success, msg = trigger_portfolio_pipeline()
        if success:
            st.success(msg)
        else:
            st.error(msg)
            
    # Expense Control
    st.subheader("📊 Expense & Cash Flow")
    if st.button("⚙️ Recalculate Budgets"):
        with st.spinner("Processing expense sheets..."):
            success, msg = trigger_expense_pipeline()
        if success:
            st.success(msg)
        else:
            st.error(msg)
            
    st.markdown("---")
    st.caption("Active Workspace Year: 2026")
    st.caption("Click the buttons above to manually sync your data.")

# Main multi-tab UI
tab_inv, tab_exp = st.tabs(["💼 Investment Portfolio", "📊 Cash Flow & Expense"])


# STREAMING_CHUNK: Formatting and displaying the Investment allocation tab...

# ==============================================================================
# TAB 1: PORTFOLIO PRESENTATION (INVESTMENTS)
# ==============================================================================
with tab_inv:
    st.header("Portfolio Asset Allocation")
    
    # Check for latest processed portfolio results
    # Using a broader wildcard to ensure we always catch the generated file
    summary_files = glob.glob(os.path.join(FINANCIAL_DATA_OUTPUT, '*Master_Account_Summary*.csv'))
    latest_summary_file = max(summary_files, key=os.path.getmtime) if summary_files else None
    
    if latest_summary_file:
        # SAFE PURE-PYTHON PARSING: Avoids memory segment faults on Python 3.14
        master_df = pd.read_csv(latest_summary_file, engine='python')
        
        # Clean currency configurations
        master_df['MarketValue'] = pd.to_numeric(master_df['MarketValue'], errors='coerce').fillna(0)
        master_df['CostBasis'] = pd.to_numeric(master_df['CostBasis'], errors='coerce').fillna(master_df['MarketValue'])
        master_df['Gain'] = master_df['MarketValue'] - master_df['CostBasis']
        master_df['AnnualIncome'] = pd.to_numeric(master_df['AnnualIncome'], errors='coerce').fillna(0)
        
        # Calculate summary indicators
        total_mv = master_df['MarketValue'].sum()
        total_gain = master_df['Gain'].sum()
        total_inc = master_df['AnnualIncome'].sum()
        
        # Render Metric Cards with color formatting
        col_m1, col_m2, col_m3 = st.columns(3)
        col_m1.metric("Total Market Value", f"${total_mv:,.0f}")
        
        # Color coding for gains
        gain_prefix = "+" if total_gain >= 0 else ""
        col_m2.metric(
            "Unrealized Gain / Loss", 
            f"{gain_prefix}${total_gain:,.0f}",
            delta=f"{total_gain/master_df['CostBasis'].sum()*100:+.2f}%" if master_df['CostBasis'].sum() > 0 else None
        )
        col_m3.metric("Annual Projected Dividends", f"${total_inc:,.0f}", delta=f"{(total_inc/total_mv)*100:.2f}% Yield" if total_mv > 0 else None)
        
        st.markdown("---")
        
        # Prepare allocation grouped layout
        overall_df = master_df.groupby('Allocation').agg({
            'MarketValue': 'sum',
            'Gain': 'sum',
            'AnnualIncome': 'sum'
        }).reset_index()
        
        overall_df['Div Rate'] = (overall_df['AnnualIncome'] / overall_df['MarketValue'] * 100).fillna(0)
        overall_df['Allocation %'] = (overall_df['MarketValue'] / total_mv * 100) if total_mv > 0 else 0.0
        total_mv = overall_df['MarketValue'].sum()
        total_gain = overall_df['Gain'].sum()

        # Append Overall Total Summary Row
        overall_total = pd.DataFrame([{
            'Allocation': 'Total Portfolio',
            'MarketValue': total_mv,
            'Gain': total_gain,
            'AnnualIncome': overall_df['AnnualIncome'].sum(),
            'Div Rate': (overall_df['AnnualIncome'].sum() / total_mv * 100) if total_mv > 0 else 0.0,
            'Allocation %': 100.0
        }])
        overall_df = pd.concat([overall_df, overall_total], ignore_index=True)

    
        st.markdown("### 📊 Performance & Composition Summary")
        
        # Create standard formatted display
        formatted_overall = overall_df.copy()
        formatted_overall['MarketValue'] = formatted_overall['MarketValue'].map('${:,.0f}'.format)
        formatted_overall['Gain'] = formatted_overall['Gain'].map('${:,.0f}'.format)
        formatted_overall['AnnualIncome'] = formatted_overall['AnnualIncome'].map('${:,.0f}'.format)
        formatted_overall['Div Rate'] = formatted_overall['Div Rate'].map('{:.2f}%'.format)
        formatted_overall['Allocation %'] = formatted_overall['Allocation %'].map('{:.2f}%'.format)
        
        # Modern use_container_width scaling prevents front-end crashes
        st.dataframe(formatted_overall, width='stretch', hide_index=True)

        st.markdown("---")
        
        # Setup charts with clean, responsive container-scaling parameters
        chart_df = overall_df.copy()
        color_encoding = alt.Color("Allocation:N", title="Asset Class", scale=alt.Scale(scheme='tableau10'))
        
        pie_value = alt.Chart(chart_df).mark_arc(innerRadius=60).encode(
            theta=alt.Theta(field="MarketValue", type="quantitative"),
            color=color_encoding,
            tooltip=[
                alt.Tooltip('Allocation:N', title='Asset Class'),
                alt.Tooltip('MarketValue:Q', title='Value ($)', format=',.0f')
            ]
        ).properties(title="Portfolio Weight (Asset Value)", height=280, width=280)

        pie_income = alt.Chart(chart_df).mark_arc(innerRadius=60).encode(
            theta=alt.Theta(field="AnnualIncome", type="quantitative"),
            color=color_encoding,
            tooltip=[
                alt.Tooltip('Allocation:N', title='Asset Class'),
                alt.Tooltip('AnnualIncome:Q', title='Value ($)', format=',.0f')
            ]
        ).properties(title="Income Contribution (Dividends)", height=280, width=280)

        # Build horizontal allocation map
        combined_charts = (pie_value | pie_income).resolve_scale(color='shared')
        
        # Place chart into container-fitting column layout
        chart_layout = st.columns([1, 0.4])
        with chart_layout[0]:
            st.altair_chart(combined_charts, width='stretch')
                    
        st.write("")
        with st.expander("🔍 Click to view detailed holding allocation sorted by Account Type"):
            acct_df = master_df.groupby(['AccountType', 'Allocation']).agg({
                'MarketValue': 'sum',
                'Gain': 'sum',
                'AnnualIncome': 'sum'
            }).reset_index()
            
            acct_df['Div Rate'] = (acct_df['AnnualIncome'] / acct_df['MarketValue'] * 100).fillna(0)
            acct_df['% of Portfolio'] = (acct_df['MarketValue'] / total_mv * 100) if total_mv > 0 else 0.0
            
            formatted_acct = acct_df.copy()
            formatted_acct['MarketValue'] = formatted_acct['MarketValue'].map('${:,.0f}'.format)
            formatted_acct['Gain'] = formatted_acct['Gain'].map('${:,.0f}'.format)
            formatted_acct['AnnualIncome'] = formatted_acct['AnnualIncome'].map('${:,.0f}'.format)
            formatted_acct['Div Rate'] = formatted_acct['Div Rate'].map('{:.2f}%'.format)
            formatted_acct['% of Portfolio'] = formatted_acct['% of Portfolio'].map('{:.2f}%'.format)
            
            st.dataframe(formatted_acct, width='stretch', hide_index=True)
            
    else:
        st.info("No Portfolio summaries found in output folder. Drop brokerage exports into 'financial_data/input/' and use sidebar to Sync.")


# ==============================================================================
# TAB 2: CASH FLOW & EXPENSE PRESENTATION
# ==============================================================================
with tab_exp:
    # 1. Load historical budget sheet numbers
    df_exp, df_inc, historical_years, tot_exp, tot_inc = load_historical_budget_data()
    
    if historical_years:
        # Create comparative timeline summary
        flow_summary = []
        for yr in historical_years:
            income_val = tot_inc.get(yr, 0.0)
            expense_val = tot_exp.get(yr, 0.0)
            net_val = income_val - expense_val
            savings_rate = (net_val / income_val * 100) if income_val > 0 else 0.0
            
            flow_summary.append({
                'Year': str(yr),
                'Total Income': income_val,
                'Total Expense': expense_val,
                'Net Cash Flow': net_val,
                'Savings Rate': savings_rate
            })
            
        summary_flow_df = pd.DataFrame(flow_summary)
        
        # Metric Overview of the most current processed year
        latest_yr_data = flow_summary[-1]
        st.subheader(f"📅 Year-to-Date Review: {latest_yr_data['Year']}")
        
        col_f1, col_f2, col_f3, col_f4 = st.columns(4)
        col_f1.metric("Annual Income", f"${latest_yr_data['Total Income']:,.0f}")
        col_f2.metric("Annual Expenses", f"${latest_yr_data['Total Expense']:,.0f}")
        
        net_flow = latest_yr_data['Net Cash Flow']
        net_prefix = "+" if net_flow >= 0 else ""
        col_f3.metric("Net Flow Amount", f"{net_prefix}${net_flow:,.0f}")
        col_f4.metric("Savings Ratio", f"{latest_yr_data['Savings Rate']:.2f}%")
        
        # Render clean categorical ledger tables
        st.markdown("---")
        st.subheader("📊 Dynamic Multi-Year Category Ledger")
        
        sub_tab_exp, sub_tab_inc = st.tabs(["💸 Expense Ledgers", "💵 Income Ledgers"])
        
        with sub_tab_exp:
            if df_exp is not None and not df_exp.empty:
                # Format Year columns
                fmt_df = df_exp.copy()
                for yr in historical_years:
                    yr_str = str(yr)
                    if yr_str in fmt_df.columns:
                        fmt_df[yr_str] = fmt_df[yr_str].map('${:,.0f}'.format)
                st.dataframe(fmt_df, width='stretch', hide_index=True)
                
        with sub_tab_inc:
            if df_inc is not None and not df_inc.empty:
                fmt_df = df_inc.copy()
                for yr in historical_years:
                    yr_str = str(yr)
                    if yr_str in fmt_df.columns:
                        fmt_df[yr_str] = fmt_df[yr_str].map('${:,.0f}'.format)
                st.dataframe(fmt_df, width='stretch', hide_index=True)

        # Convert list of summaries for side-by-side charting
        st.markdown("---")                
        
        flow_chart_df = pd.DataFrame({
            'Year': [s['Year'] for s in flow_summary],
            'Income': [s['Total Income'] for s in flow_summary],
            'Expense': [s['Total Expense'] for s in flow_summary],
            'Savings': [s['Net Cash Flow'] for s in flow_summary]
        })
        
        # Render historical flows
        hist_cols = st.columns([1, 1])
        
        with hist_cols[0]:
            st.markdown("#### Historical Cash Flow Trend (Last 5 Years)")
            melted_df = pd.melt(flow_chart_df, id_vars=['Year'], value_vars=['Income', 'Expense'], var_name='Type', value_name='Amount')
            
            bar_chart = alt.Chart(melted_df).mark_bar().encode(
                x=alt.X('Year:N', title='Budget Year'),
                y=alt.Y('Amount:Q', title='Value ($)'),
                color=alt.Color('Type:N', scale=alt.Scale(domain=['Income', 'Expense'], range=['#2ca02c', '#d62728'])),
                xOffset='Type:N'
            ).properties(height=280, width=420)
            
            st.altair_chart(bar_chart, width='stretch')
            
        with hist_cols[1]:
            st.markdown("#### Annual Net Wealth Savings Rate")
            line_chart = alt.Chart(flow_chart_df).mark_line(point=True, color='#0d6efd').encode(
                x=alt.X('Year:N', title='Budget Year'),
                y=alt.Y('Savings:Q', title='Savings Generated ($)'),
                tooltip=['Year:N', alt.Tooltip('Savings:Q', format=',.2f')]
            ).properties(height=280, width=420)
            
            st.altair_chart(line_chart, width='stretch')
                    
        # Alerts & Action items block
        alerts_data = parse_alerts_log()
        
        if alerts_data and (alerts_data['unmapped'] or alerts_data['others'] or alerts_data['abnormal']):
            st.markdown("### 🚨 Transaction Alerts & Operations Center")
            st.caption(f"Log generated from run file dating: {alerts_data.get('file_date', 'N/A')}")
            
            # Unmapped warnings
            if alerts_data['unmapped']:
                with st.expander("⚠️ Unmapped Transaction Categories Found"):                
                    st.markdown(
                        f"""<div class="alert-card alert-danger">
                        <strong>⚠️ Unmapped Transaction Categories Found:</strong><br/>
                        These categories exist in your statements but are missing in <code>category_type_mapping.csv</code>.
                        They are tagged <strong>UNMAPPED</strong>. Please update mapping table and recalculate.
                        <ul style="margin-top: 5px; margin-bottom: 0;">
                            {"".join([f"<li>{cat}</li>" for cat in alerts_data['unmapped']])}
                        </ul>
                        </div>""", unsafe_allow_html=True
                    )
                
            # Category others rollover warnings
            if alerts_data['others']:
                with st.expander("📋 Categories Rolled into 'Others'"):
                    st.markdown(
                        f"""<div class="alert-card alert-warning">
                        <strong>📋 Categories Rolled into 'Others':</strong><br/>
                        These budget categories don't have matching line items in <code>Yearly_budget.xlsx</code> 
                        and lack an alias mapping configuration, defaulting to 'Others'. Create row entries or custom aliases.
                        <ul style="margin-top: 5px; margin-bottom: 0;">
                            {"".join([f"<li>{cat}</li>" for cat in alerts_data['others']])}
                        </ul>
                        </div>""", unsafe_allow_html=True
                    )
                
            # Outliers detector
            if alerts_data['abnormal']:
                with st.expander("🔎 Unusual / Abnormal Transaction Outliers Flagged"):
                    st.markdown(
                        f"""<div class="alert-card alert-warning">
                        <strong>🔎 Unusual / Abnormal Transaction Outliers Flagged</strong><br/>
                        These transactions appear to be outliers and are flagged to study further.
                        <ul style="margin-top: 5px; margin-bottom: 0;">
                            {"".join([f"<li>{cat}</li>" for cat in alerts_data['abnormal']])}
                        </ul>
                        </div>""", unsafe_allow_html=True
                    )
        else:
            st.success("✨ Zero unmapped categories or extreme transaction anomalies flagged this cycle!")
            

    else:
        st.info("No transaction spreadsheets found in outputs. Drop transactions CSV and budget templates into 'expense_data/input/' to start calculation workflows.")
