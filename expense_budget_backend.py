"""
Expense/Income/Transfer tracking automation.

Workflow this replaces (previously done by hand each time in Excel):
  1. Download a YTD transactions CSV from Empower Personal Dashboard.
  2. Tag each row Expense / Income / Transfer based on its Category (VLOOKUP against
     a Category -> Type lookup table).
  3. Build a pivot-style summary: Type > Category, by month.
  4. Copy that year's category totals into Yearly_budget.xlsx.

Run it each time you download a fresh export:
    python expense_budget_backend.py

Directory layout (created automatically if missing):
    expense_data/
        input/    <- drop the raw Empower CSV export and Yearly_budget.xlsx here
                  <- category_type_mapping.csv, budget_category_aliases.csv (edit these directly)
        output/   <- {year}_transactions_processed.xlsx and updated Yearly_budget.xlsx land here
"""
import glob
import os
import re
import sys
from datetime import datetime

from copy import copy
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

BASE_DIR = 'expense_data'
INPUT_DIR = os.path.join(BASE_DIR, 'input')
CONFIG_DIR = os.path.join(BASE_DIR, 'input')
OUTPUT_DIR = os.path.join(BASE_DIR, 'output')

CATEGORY_TYPE_MAP_PATH = os.path.join(CONFIG_DIR, 'category_type_mapping.csv')
BUDGET_ALIAS_MAP_PATH = os.path.join(CONFIG_DIR, 'budget_category_aliases.csv')

MONTHS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

# Outlier detection: a transaction is flagged if its |Amount| is unusually
# large for its Category, among Expense/Income transactions only (Transfer and
# investment-trade categories naturally swing to large dollar amounts and
# aren't meaningful to flag here). Categories need at least MIN_CATEGORY_COUNT
# transactions to get a category-specific bound (a multiple of that category's
# own median); thinner categories fall back to a flat dollar threshold since
# there isn't enough history to know what's "normal" there.
ABNORMAL_TYPES = ('Expense', 'Income')
MEDIAN_MULTIPLIER = 8.0
MIN_CATEGORY_COUNT = 5
FALLBACK_ABS_THRESHOLD = 2000.0


# ---------------------------------------------------------------------------
# Setup / config helpers
# ---------------------------------------------------------------------------

def ensure_dirs():
    for d in (INPUT_DIR, CONFIG_DIR, OUTPUT_DIR):
        os.makedirs(d, exist_ok=True)


def load_category_type_map():
    """Category -> Type (Expense/Income/Transfer). Edit the CSV directly to add new categories."""
    if not os.path.exists(CATEGORY_TYPE_MAP_PATH):
        pd.DataFrame(columns=['Category', 'Type']).to_csv(CATEGORY_TYPE_MAP_PATH, index=False)
    df = pd.read_csv(CATEGORY_TYPE_MAP_PATH)
    return dict(zip(df['Category'].str.strip(), df['Type'].str.strip()))


def load_budget_aliases():
    """Category -> Yearly_budget row label, only needed when the names differ.
    Anything not listed here either matches a budget row by name directly, or
    falls into that sheet's 'Others' catch-all row."""
    if not os.path.exists(BUDGET_ALIAS_MAP_PATH):
        pd.DataFrame(columns=['Category', 'BudgetRow']).to_csv(BUDGET_ALIAS_MAP_PATH, index=False)
    df = pd.read_csv(BUDGET_ALIAS_MAP_PATH)
    return dict(zip(df['Category'].str.strip(), df['BudgetRow'].str.strip()))


def find_input_file(patterns):
    matches = []
    for pat in patterns:
        matches.extend(glob.glob(os.path.join(INPUT_DIR, pat)))
    if not matches:
        return None
    return max(matches, key=os.path.getmtime)


# ---------------------------------------------------------------------------
# Step 1: read the raw Empower export, tag each row, infer the target year
# ---------------------------------------------------------------------------

def load_raw_transactions(csv_path, category_type_map):
    df = pd.read_csv(csv_path)
    df['Date'] = pd.to_datetime(df['Date'])
    df['Category'] = df['Category'].astype(str).str.strip()
    df['Type'] = df['Category'].map(category_type_map).fillna('UNMAPPED')

    unmapped_alerts = []
    unmapped = sorted(df.loc[df['Type'] == 'UNMAPPED', 'Category'].unique())
    if unmapped:
        unmapped_total = df.loc[df['Type'] == 'UNMAPPED', 'Amount'].sum()
        print(f"WARNING: {len(unmapped)} categor{'y is' if len(unmapped) == 1 else 'ies are'} "
              f"not in {CATEGORY_TYPE_MAP_PATH}, totaling ${unmapped_total:,.2f}:")
        for cat in unmapped:
            cat_total = df.loc[df['Category'] == cat, 'Amount'].sum()
            cat_count = int((df['Category'] == cat).sum())
            print(f"    - {cat}  (${cat_total:,.2f})")
            unmapped_alerts.append({'Category': cat, 'Count': cat_count, 'Total': cat_total})
        print(f"  Add these to {CATEGORY_TYPE_MAP_PATH} (Category,Type) and re-run to include "
              f"them as Expense/Income/Transfer. For now they're tagged UNMAPPED.")

    return df, unmapped_alerts


def detect_abnormal_transactions(df):
    """Flags Expense/Income transactions whose |Amount| is unusually large for
    their Category (a multiple of that category's own median, or a flat dollar
    floor for thin categories). Transfer/investment-trade categories are
    excluded since large dollar swings are routine there, not abnormal.
    """
    flagged_rows = []
    scoped = df.loc[df['Type'].isin(ABNORMAL_TYPES)]
    abs_amount = scoped['Amount'].abs()

    for cat, group in scoped.groupby('Category'):
        idx = group.index
        amounts = abs_amount.loc[idx]
        if len(group) >= MIN_CATEGORY_COUNT:
            bound = max(amounts.median() * MEDIAN_MULTIPLIER, FALLBACK_ABS_THRESHOLD)
        else:
            bound = FALLBACK_ABS_THRESHOLD

        over = idx[amounts > bound]
        for i in over:
            flagged_rows.append({
                'Date': df.loc[i, 'Date'],
                'Account': df.loc[i, 'Account'],
                'Description': df.loc[i, 'Description'],
                'Category': cat,
                'Amount': df.loc[i, 'Amount'],
                'Reason': f'|Amount| exceeds normal range for "{cat}" (threshold ${bound:,.0f})',
            })

    flagged_df = pd.DataFrame(flagged_rows)
    if not flagged_df.empty:
        flagged_df = flagged_df.sort_values('Amount', key=lambda s: s.abs(), ascending=False)
    return flagged_df

def infer_target_year(csv_path, df):
    # Try the "YYYY-MM-DD_thru_YYYY-MM-DD" filename convention first
    fname = os.path.basename(csv_path)
    m = re.search(r'(\d{4})-\d{2}-\d{2}_thru_\d{4}-\d{2}-\d{2}', fname)
    if m:
        return int(m.group(1))
    # Fall back to the most common year in the data
    return int(df['Date'].dt.year.mode().iloc[0])


# ---------------------------------------------------------------------------
# Step 2: write the per-year transactions workbook (Sheet1 + Lookup + Summary)
# ---------------------------------------------------------------------------

def write_transactions_workbook(df, category_type_map, year, out_path, unmapped_alerts,
                                 abnormal_df, others_alerts):
    wb = Workbook()

    # --- Lookup sheet (Category -> Type), written first since Sheet1's Tags
    # column formula references it ---
    ws_lookup = wb.active
    ws_lookup.title = 'Lookup'
    ws_lookup['A1'] = 'Category'
    ws_lookup['B1'] = 'Type'
    for i, (cat, typ) in enumerate(sorted(category_type_map.items()), start=2):
        ws_lookup.cell(row=i, column=1, value=cat)
        ws_lookup.cell(row=i, column=2, value=typ)
    last_lookup_row = len(category_type_map) + 1

    # --- Sheet1: raw transactions, Tags column re-created as a live formula
    # so it matches the original workbook's convention and recalculates if you
    # ever hand-edit a Category cell ---
    ws_raw = wb.create_sheet('Sheet1')
    headers = ['Date', 'Account', 'Description', 'Category', 'Tags', 'Amount']
    ws_raw.append(headers)
    for cell in ws_raw[1]:
        cell.font = Font(bold=True)

    df_sorted = df.sort_values('Date').reset_index(drop=True)
    for i, row in enumerate(df_sorted.itertuples(index=False), start=2):
        ws_raw.cell(row=i, column=1, value=row.Date.to_pydatetime()).number_format = 'yyyy-mm-dd'
        ws_raw.cell(row=i, column=2, value=row.Account)
        ws_raw.cell(row=i, column=3, value=row.Description)
        ws_raw.cell(row=i, column=4, value=row.Category)
        ws_raw.cell(row=i, column=5, value=(
            f'=IFERROR(VLOOKUP(D{i}, Lookup!A$2:B${last_lookup_row}, 2, FALSE), "UNMAPPED")'
        ))
        ws_raw.cell(row=i, column=6, value=row.Amount)
    last_data_row = len(df_sorted) + 1

    # --- Summary sheet: Type > Category, by month (mirrors the old PivotTable) ---
    ws_sum = wb.create_sheet('Summary')
    ws_sum['A1'] = 'Row Labels'
    for j, mon in enumerate(MONTHS, start=2):
        ws_sum.cell(row=1, column=j, value=mon)
    ws_sum.cell(row=1, column=14, value='Grand Total')
    for cell in ws_sum[1]:
        cell.font = Font(bold=True)

    # Order: Expense, Income, Transfer, then anything else (e.g. UNMAPPED) that showed up
    present_types = list(df['Type'].unique())
    type_order = [t for t in ['Expense', 'Income', 'Transfer'] if t in present_types]
    type_order += sorted(t for t in present_types if t not in type_order)

    r = 2
    type_subtotal_rows = []
    accounting_fmt = '"$"#,##0_);("$"#,##0)'
    
    for typ in type_order:
        cats_in_type = df.loc[df['Type'] == typ]
        # order categories by total magnitude, largest first, like the original pivot
        cat_order = (cats_in_type.groupby('Category')['Amount'].sum()
                     .abs().sort_values(ascending=False).index.tolist())

        type_row = r
        type_subtotal_rows.append(type_row)
        ws_sum.cell(row=type_row, column=1, value=typ).font = Font(bold=True)
        r += 1
        first_cat_row = r
        for cat in cat_order:
            ws_sum.cell(row=r, column=1, value=cat)
            ws_sum.row_dimensions[r].outline_level = 1
            for j, _mon in enumerate(MONTHS, start=2):
                m = j - 1
                col_letter = get_column_letter(6)  # Amount is column F on Sheet1
                cat_escaped = cat.replace('"', '""')
                formula = (
                    f'=SUMIFS(Sheet1!${col_letter}$2:${col_letter}${last_data_row}, '
                    f'Sheet1!$D$2:$D${last_data_row}, "{cat_escaped}", '
                    f'Sheet1!$A$2:$A${last_data_row}, ">="&DATE({year},{m},1), '
                    f'Sheet1!$A$2:$A${last_data_row}, "<="&EOMONTH(DATE({year},{m},1),0))'
                )
                ws_sum.cell(row=r, column=j, value=formula).number_format = accounting_fmt
            ws_sum.cell(row=r, column=14, value=f'=SUM(B{r}:M{r})').number_format = accounting_fmt
            r += 1
        last_cat_row = r - 1
        # Type subtotal = sum of its category rows
        for j in range(2, 15):
            col = get_column_letter(j)
            ws_sum.cell(row=type_row, column=j,
                        value=f'=SUM({col}{first_cat_row}:{col}{last_cat_row})').number_format = accounting_fmt

    grand_total_row = r
    ws_sum.cell(row=grand_total_row, column=1, value='Grand Total').font = Font(bold=True)
    for j in range(2, 15):
        col = get_column_letter(j)
        refs = ','.join(f'{col}{tr}' for tr in type_subtotal_rows)
        ws_sum.cell(row=grand_total_row, column=j, value=f'=SUM({refs})').number_format = accounting_fmt

    # --- Formatting Widths ---
    for ws in (ws_lookup, ws_raw):
        for col_cells in ws.columns:
            length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 45)

    # Specific logic for Summary tab to keep columns readable
    for i, col_cells in enumerate(ws_sum.columns, start=1):
        letter = get_column_letter(i)
        if i == 1: # Column A: Row Labels (Auto-fit)
            length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
            ws_sum.column_dimensions[letter].width = min(max(length + 2, 10), 30)
        else: # Columns B-N: Months and Totals (Fixed width)
            ws_sum.column_dimensions[letter].width = 13

    write_alerts_sheet(wb, unmapped_alerts, abnormal_df, others_alerts)
    wb.save(out_path)
    return {
        'type_subtotal_rows': dict(zip(type_order, type_subtotal_rows)),
        'grand_total_row': grand_total_row,
    }


# ---------------------------------------------------------------------------
# Step 3: roll category totals up to Yearly_budget.xlsx rows, update that year's column
# ---------------------------------------------------------------------------

def compute_budget_rollup(df, sheet_type, budget_row_labels, aliases):
    """Returns ({budget_row_label: total}, unmatched) for one Type (Expense or
    Income), using the sign convention Yearly_budget.xlsx already uses: Expense
    totals are stored as positive spend magnitudes, Income totals stay as-is."""
    sub = df.loc[df['Type'] == sheet_type]
    normalized_labels = {lbl.strip().lower(): lbl for lbl in budget_row_labels}

    rollup = {lbl: 0.0 for lbl in budget_row_labels}
    unmatched = []

    cat_totals = sub.groupby('Category')['Amount'].sum()
    for cat, total in cat_totals.items():
        target = aliases.get(cat)
        if target is None:
            target = normalized_labels.get(cat.strip().lower())
        if target is None:
            target = 'Others' if 'Others' in budget_row_labels else None
            if target is not None:
                unmatched.append((cat, total))
        if target is None:
            unmatched.append((cat, total))
            continue
        signed_total = -total if sheet_type == 'Expense' else total
        rollup[target] = rollup.get(target, 0.0) + signed_total

    if unmatched:
        print(f"NOTE: {sheet_type} categories rolled into 'Others' on the {sheet_type} budget "
              f"sheet (no matching row and no alias defined):")
        for cat, total in unmatched:
            print(f"    - {cat}  (${total:,.2f})")
        print(f"  Add an alias to {BUDGET_ALIAS_MAP_PATH} if one of these deserves its own row.")

    return rollup, unmatched

def copy_cell_style(source_cell, target_cell):
    """
    Copies all styling attributes from a source cell to a target cell.
    Uses copy() to prevent shared reference bugs in openpyxl.
    """
    if source_cell and source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = source_cell.number_format
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.protection = copy(source_cell.protection)

def update_budget_sheet(ws, year, rollup):
    header_row = 1
    
    # 1. Identify existing year column or prepare to create one
    last_header_col = 1
    for col in range(2, ws.max_column + 1):
        if str(ws.cell(row=header_row, column=col).value or '').strip():
            last_header_col = col
            
    year_col = None
    for col in range(2, last_header_col + 1):
        header_val = str(ws.cell(row=header_row, column=col).value or '').strip()
        if header_val == str(year):
            year_col = col
            break
            
    # Add new year column if missing
    if year_col is None:
        year_col = last_header_col + 1
        header_cell = ws.cell(row=header_row, column=year_col, value=str(year))
        # Copy style from the header immediately to the left
        if year_col > 1:
            copy_cell_style(ws.cell(row=header_row, column=year_col - 1), header_cell)

    # 2. Robustly find the 'Total' row
    total_row = None
    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val and str(val).strip().lower() == 'total':
            total_row = r
            break
    
    if total_row is None:
        # Fallback: if 'Total' row not found, cannot proceed safely
        raise ValueError("Critical error: Could not locate a row labeled 'Total' in the sheet.")
        
    last_cat_row = total_row - 1

    # 3. Populate categories, clone styles from the previous year's column
    source_col = year_col - 1
    # Only copy style if we have a valid previous year column to reference
    should_copy_style = (source_col >= 2) 

    for row in range(2, last_cat_row + 1):
        label = str(ws.cell(row=row, column=1).value or '').strip()
        match = next((k for k in rollup if k.strip().lower() == label.lower()), None)
        value = rollup.get(match, 0.0) if match else 0.0
        
        target_cell = ws.cell(row=row, column=year_col, value=round(value, 2))
        
        if should_copy_style:
            copy_cell_style(ws.cell(row=row, column=source_col), target_cell)

    # 4. Write total formula and apply bold style
    col_letter = get_column_letter(year_col)
    total_cell = ws.cell(row=total_row, column=year_col,
                         value=f'=SUM({col_letter}2:{col_letter}{last_cat_row})')
                         
    if should_copy_style:
        copy_cell_style(ws.cell(row=total_row, column=source_col), total_cell)
        
    ws.cell(row=total_row, column=1).font = Font(bold=True)


def update_yearly_budget(budget_path, year, df, aliases, out_path):
    wb = load_workbook(budget_path)
    others_alerts = {}

    ws_exp = wb['Expense']
    exp_labels = [str(ws_exp.cell(row=r, column=1).value).strip()
                  for r in range(2, ws_exp.max_row + 1)
                  if ws_exp.cell(row=r, column=1).value
                  and str(ws_exp.cell(row=r, column=1).value).strip().lower() != 'total']
    exp_rollup, exp_unmatched = compute_budget_rollup(df, 'Expense', exp_labels, aliases)
    update_budget_sheet(ws_exp, year, exp_rollup)
    others_alerts['Expense'] = exp_unmatched

    ws_inc = wb['Income']
    inc_labels = [str(ws_inc.cell(row=r, column=1).value).strip()
                  for r in range(2, ws_inc.max_row + 1)
                  if ws_inc.cell(row=r, column=1).value
                  and str(ws_inc.cell(row=r, column=1).value).strip().lower() != 'total']
    inc_rollup, inc_unmatched = compute_budget_rollup(df, 'Income', inc_labels, aliases)
    update_budget_sheet(ws_inc, year, inc_rollup)
    others_alerts['Income'] = inc_unmatched

    wb.save(out_path)
    return others_alerts


# ---------------------------------------------------------------------------
# Alerts: one sheet in the transactions workbook + a plaintext log, so issues
# are visible whether you open Excel or just skim the terminal/output folder.
# ---------------------------------------------------------------------------

def write_alerts_sheet(wb, unmapped_alerts, abnormal_df, others_alerts):
    ws = wb.create_sheet('Alerts')
    r = 1

    def section(title):
        nonlocal r
        ws.cell(row=r, column=1, value=title).font = Font(bold=True, size=12)
        r += 1

    def subheader(*labels):
        nonlocal r
        for j, label in enumerate(labels, start=1):
            ws.cell(row=r, column=j, value=label).font = Font(bold=True)
        r += 1

    any_alerts = bool(unmapped_alerts) or not abnormal_df.empty or any(others_alerts.values())
    if not any_alerts:
        ws.cell(row=r, column=1, value='No alerts - nothing unmapped or abnormal this run.')
        r += 1

    if unmapped_alerts:
        section('Unmapped categories (not in category_type_mapping.csv - tagged UNMAPPED)')
        subheader('Category', 'Count', 'Total Amount')
        for a in unmapped_alerts:
            ws.cell(row=r, column=1, value=a['Category'])
            ws.cell(row=r, column=2, value=a['Count'])
            ws.cell(row=r, column=3, value=round(a['Total'], 2))
            r += 1
        r += 1

    for sheet_type, unmatched in others_alerts.items():
        if not unmatched:
            continue
        section(f"{sheet_type} categories rolled into 'Others' on the {sheet_type} budget sheet")
        subheader('Category', 'Total Amount')
        for cat, total in unmatched:
            ws.cell(row=r, column=1, value=cat)
            ws.cell(row=r, column=2, value=round(total, 2))
            r += 1
        r += 1

    if not abnormal_df.empty:
        section('Abnormal transactions (unusually large for their category)')
        subheader('Date', 'Account', 'Description', 'Category', 'Amount', 'Reason')
        for _, row in abnormal_df.iterrows():
            ws.cell(row=r, column=1, value=row['Date'].strftime('%Y-%m-%d'))
            ws.cell(row=r, column=2, value=row['Account'])
            ws.cell(row=r, column=3, value=row['Description'])
            ws.cell(row=r, column=4, value=row['Category'])
            ws.cell(row=r, column=5, value=round(row['Amount'], 2))
            ws.cell(row=r, column=6, value=row['Reason'])
            r += 1

    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 60)


def write_alerts_log(log_path, unmapped_alerts, abnormal_df, others_alerts):
    lines = [f"Alerts - generated {datetime.now().strftime('%Y-%m-%d %H:%M')}", '=' * 60, '']

    if unmapped_alerts:
        lines.append('UNMAPPED CATEGORIES (add to category_type_mapping.csv):')
        for a in unmapped_alerts:
            lines.append(f"  - {a['Category']}  ({a['Count']} txns, ${a['Total']:,.2f})")
        lines.append('')

    for sheet_type, unmatched in others_alerts.items():
        if not unmatched:
            continue
        lines.append(f"{sheet_type.upper()} CATEGORIES ROLLED INTO 'OTHERS' "
                     f"(add an alias in budget_category_aliases.csv if one deserves its own row):")
        for cat, total in unmatched:
            lines.append(f"  - {cat}  (${total:,.2f})")
        lines.append('')

    if not abnormal_df.empty:
        lines.append(f'ABNORMAL TRANSACTIONS ({len(abnormal_df)} flagged):')
        for _, row in abnormal_df.iterrows():
            lines.append(f"  - {row['Date'].strftime('%Y-%m-%d')}  {row['Category']:<25} "
                         f"${row['Amount']:>12,.2f}  {row['Description']}  ({row['Account']})")
        lines.append('')

    if len(lines) == 3:
        lines.append('No alerts - nothing unmapped or abnormal this run.')

    with open(log_path, 'w') as f:
        f.write('\n'.join(lines))


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def run():
    ensure_dirs()

    csv_path = find_input_file(['*_thru_*_transactions.csv', '*transactions*.csv'])
    if not csv_path:
        print(f"No transactions CSV found in {INPUT_DIR}/. Drop your Empower export there and re-run.")
        sys.exit(1)
    print(f"Using transactions file: {csv_path}")

    budget_path = find_input_file(['Yearly_budget.xlsx', '*budget*.xlsx'])
    if not budget_path:
        print(f"No Yearly_budget.xlsx found in {INPUT_DIR}/. Skipping the budget update step.")

    category_type_map = load_category_type_map()
    aliases = load_budget_aliases()

    df, unmapped_alerts = load_raw_transactions(csv_path, category_type_map)
    year = infer_target_year(csv_path, df)
    print(f"Target year: {year}  ({len(df)} transactions)")

    abnormal_df = detect_abnormal_transactions(df)
    if not abnormal_df.empty:
        print(f"WARNING: {len(abnormal_df)} transaction(s) look abnormally large for their category:")
        for _, row in abnormal_df.iterrows():
            print(f"    - {row['Date'].strftime('%Y-%m-%d')}  {row['Category']}  "
                  f"${row['Amount']:,.2f}  ({row['Description']})")

    others_alerts = {}
    budget_out_path = os.path.join(OUTPUT_DIR, 'Yearly_budget.xlsx')
    if budget_path:
        others_alerts = update_yearly_budget(budget_path, year, df, aliases, budget_out_path)

    txn_out_path = os.path.join(OUTPUT_DIR, f'{year}_transactions_processed.xlsx')
    write_transactions_workbook(df, category_type_map, year, txn_out_path,
                                unmapped_alerts, abnormal_df, others_alerts)
    print(f"Wrote transactions workbook: {txn_out_path}")

    log_path = os.path.join(OUTPUT_DIR, f'{year}_alerts.txt')
    write_alerts_log(log_path, unmapped_alerts, abnormal_df, others_alerts)
    print(f"Wrote alerts log: {log_path}")

    if budget_path:
        print(f"Wrote updated budget workbook: {budget_out_path}")
        print("Review the output copy, then replace your working Yearly_budget.xlsx with it "
              "once you're happy with the numbers.")


if __name__ == '__main__':
    run()